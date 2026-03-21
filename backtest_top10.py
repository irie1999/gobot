"""
バックテスト: swing_notify_top10.py 戦略
過去5年の日足データで10銘柄一括検証（S株 1株単位）

使い方:
  python backtest_top10.py
"""

import os
import subprocess
import sys
import warnings
warnings.filterwarnings("ignore")

from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter

# ── 戦略パラメータ（swing_notify_top10.py と完全一致） ──────────
WATCH_LIST = [
    ("8604.T", "野村HD"),
    ("2802.T", "味の素"),
    ("2914.T", "JT"),
    ("7203.T", "トヨタ自動車"),
    ("8001.T", "伊藤忠商事"),
    ("8306.T", "三菱UFJ"),
    ("6752.T", "パナソニックHD"),
    ("8002.T", "丸紅"),
    ("9020.T", "JR東日本"),
    ("6902.T", "デンソー"),
]

EMA_FAST       = 5
EMA_MID        = 20
EMA_SLOW       = 50
RSI_PERIOD     = 14
RSI_ENTRY      = 55
RSI_EXIT       = 60
BB_PERIOD      = 20
BB_K           = 2.0
ATR_PERIOD     = 14
ATR_STOP_MULT  = 1.5
RISK_PER_TRADE = 0.03
INITIAL_CASH   = 500_000
MAX_COST_RATIO = 0.10
MAX_QTY        = 3000


# ── データ取得 ────────────────────────────────────────────────
def fetch_data(symbol: str, start: str, end: str) -> pd.DataFrame:
    df = yf.download(symbol, start=start, end=end, progress=False, auto_adjust=True)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.index = pd.to_datetime(df.index)
    if df.index.tz is not None:
        df.index = df.index.tz_convert("Asia/Tokyo").tz_localize(None)
    df.columns = [c.lower() for c in df.columns]
    return df[["open", "high", "low", "close", "volume"]].dropna()


# ── インジケーター（swing_notify_top10.py と同一） ────────────
def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    c = df["close"]
    h = df["high"]
    l = df["low"]

    df["ema_fast"] = c.ewm(span=EMA_FAST,  adjust=False).mean()
    df["ema_mid"]  = c.ewm(span=EMA_MID,   adjust=False).mean()
    df["ema_slow"] = c.ewm(span=EMA_SLOW,  adjust=False).mean()
    df["ema_cross_up"] = (
        (df["ema_fast"] > df["ema_mid"]) &
        (df["ema_fast"].shift(1) <= df["ema_mid"].shift(1))
    )

    delta = c.diff()
    gain  = delta.clip(lower=0).ewm(com=RSI_PERIOD - 1, adjust=False).mean()
    loss  = (-delta.clip(upper=0)).ewm(com=RSI_PERIOD - 1, adjust=False).mean()
    df["rsi"] = 100 - (100 / (1 + gain / loss.replace(0, np.nan))).fillna(100)

    sma = c.rolling(BB_PERIOD).mean()
    std = c.rolling(BB_PERIOD).std(ddof=0)
    df["bb_upper"] = sma + BB_K * std
    df["bb_lower"] = sma - BB_K * std
    df["bb_band"]  = BB_K * std

    prev_c = c.shift(1)
    tr = pd.concat([h - l,
                    (h - prev_c).abs(),
                    (l - prev_c).abs()], axis=1).max(axis=1)
    df["atr"] = tr.ewm(span=ATR_PERIOD, adjust=False).mean()
    return df


# ── シグナル（swing_notify_top10.py と同一） ─────────────────
def add_signals(df: pd.DataFrame) -> pd.DataFrame:
    trend_up = df["close"] > df["ema_slow"]
    df["entry_sig"] = trend_up & (
        (df["rsi"] < RSI_ENTRY) |
        df["ema_cross_up"] |
        (df["close"] <= df["bb_lower"] + df["bb_band"])
    )
    df["exit_sig"] = (
        (df["rsi"]      > RSI_EXIT) |
        (df["close"]    >= df["bb_upper"]) |
        (df["ema_fast"] < df["ema_mid"])
    )
    return df


# ── 株数計算（swing_notify_top10.py と同一） ─────────────────
def calc_qty(cash: float, atr: float, close: float) -> int:
    stop_dist = atr * ATR_STOP_MULT
    if stop_dist <= 0:
        return 0
    qty_by_risk = int(cash * RISK_PER_TRADE / stop_dist)
    qty_by_cost = int(cash * MAX_COST_RATIO / close) if close > 0 else qty_by_risk
    qty = min(qty_by_risk, qty_by_cost, MAX_QTY)
    return max(qty, 1)


# ── 1銘柄バックテスト ─────────────────────────────────────────
def backtest_symbol(df: pd.DataFrame, symbol: str, name: str) -> dict:
    """
    シグナルは前日終値で確認し、翌日始値で執行。
    ストップは当日の安値がストップ価格以下になった場合に執行
    （始値がギャップダウンしていれば始値で決済）。
    """
    cash = INITIAL_CASH
    in_pos = False
    entry_price = stop_price = 0.0
    qty = 0
    entry_dt = None
    trades = []

    for i in range(1, len(df)):
        today = df.iloc[i]
        prev  = df.iloc[i - 1]

        # ── 保有中: 決済判定 ─────────────────────────────────
        if in_pos:
            exit_price  = None
            exit_reason = None

            if today["low"] <= stop_price:
                # ストップ到達: ギャップダウン考慮
                exit_price  = min(today["open"], stop_price)
                exit_reason = "stop"
            elif prev["exit_sig"]:
                exit_price  = today["open"]
                exit_reason = "signal"

            if exit_price is not None:
                pnl    = (exit_price - entry_price) * qty
                cash  += exit_price * qty
                trades.append({
                    "entry_dt":    entry_dt,
                    "exit_dt":     df.index[i],
                    "symbol":      symbol,
                    "name":        name,
                    "entry_price": entry_price,
                    "exit_price":  exit_price,
                    "qty":         qty,
                    "pnl":         pnl,
                    "reason":      exit_reason,
                })
                in_pos = False

        # ── 待機中: エントリー判定 ───────────────────────────
        if not in_pos and prev["entry_sig"]:
            q    = calc_qty(cash, prev["atr"], prev["close"])
            cost = today["open"] * q
            if q > 0 and cost <= cash:
                entry_price = today["open"]
                stop_price  = entry_price - prev["atr"] * ATR_STOP_MULT
                qty         = q
                cash       -= cost
                entry_dt    = df.index[i]
                in_pos      = True

    # 未決済: 最終日終値で強制クローズ
    if in_pos:
        exit_price = df.iloc[-1]["close"]
        pnl        = (exit_price - entry_price) * qty
        cash      += exit_price * qty
        trades.append({
            "entry_dt":    entry_dt,
            "exit_dt":     df.index[-1],
            "symbol":      symbol,
            "name":        name,
            "entry_price": entry_price,
            "exit_price":  exit_price,
            "qty":         qty,
            "pnl":         pnl,
            "reason":      "force_close",
        })

    total_pnl = cash - INITIAL_CASH
    n_trades  = len(trades)

    wins   = [t["pnl"] for t in trades if t["pnl"] > 0]
    losses = [t["pnl"] for t in trades if t["pnl"] <= 0]
    pf     = sum(wins) / abs(sum(losses)) if losses and sum(losses) != 0 else float("inf")

    return {
        "symbol":        symbol,
        "name":          name,
        "trades":        n_trades,
        "win_rate":      len(wins) / n_trades * 100 if n_trades else 0.0,
        "total_pnl":     total_pnl,
        "return_pct":    total_pnl / INITIAL_CASH * 100,
        "avg_pnl":       np.mean([t["pnl"] for t in trades]) if trades else 0.0,
        "max_win":       max(wins)    if wins   else 0.0,
        "max_loss":      min(losses)  if losses else 0.0,
        "profit_factor": pf,
        "trades_detail": trades,
    }


# ── ポートフォリオ指標 ────────────────────────────────────────
def portfolio_metrics(all_trades: list) -> dict:
    if not all_trades:
        return {}

    td = pd.DataFrame(all_trades).sort_values("exit_dt").reset_index(drop=True)

    # 累積損益 → 最大ドローダウン
    td["cum_pnl"] = td["pnl"].cumsum()
    initial_total = INITIAL_CASH * len(WATCH_LIST)
    equity        = initial_total + td["cum_pnl"]
    peak          = equity.cummax()
    drawdown      = (peak - equity) / peak * 100
    max_dd        = drawdown.max()

    wins   = td[td["pnl"] > 0]
    losses = td[td["pnl"] <= 0]
    pf     = wins["pnl"].sum() / abs(losses["pnl"].sum()) \
             if len(losses) > 0 and losses["pnl"].sum() != 0 else float("inf")

    td["hold_days"] = (td["exit_dt"] - td["entry_dt"]).dt.days

    return {
        "n_trades":      len(td),
        "win_rate":      len(wins) / len(td) * 100,
        "profit_factor": pf,
        "max_drawdown":  max_dd,
        "avg_hold_days": td["hold_days"].mean(),
        "avg_pnl":       td["pnl"].mean(),
        "best_trade":    td["pnl"].max(),
        "worst_trade":   td["pnl"].min(),
        "stop_count":    len(td[td["reason"] == "stop"]),
        "signal_count":  len(td[td["reason"] == "signal"]),
        "force_count":   len(td[td["reason"] == "force_close"]),
    }


# ── Excel エクスポート ────────────────────────────────────────
def export_excel(results: list, all_trades: list, m: dict,
                 start: str, end: str, path: str = "backtest_result.xlsx") -> str:

    total_pnl  = sum(r["total_pnl"] for r in results)
    total_init = INITIAL_CASH * len(WATCH_LIST)
    total_ret  = total_pnl / total_init * 100
    years      = (datetime.strptime(end, "%Y-%m-%d") -
                  datetime.strptime(start, "%Y-%m-%d")).days / 365.25
    cagr       = ((total_init + total_pnl) / total_init) ** (1 / years) - 1

    wb = Workbook()

    # ── 共通スタイル定義 ─────────────────────────────────────
    C_HEADER_BG  = "1F3864"   # 濃紺
    C_HEADER_FG  = "FFFFFF"
    C_SUB_BG     = "2E75B6"   # 中青
    C_ALT        = "D9E1F2"   # 薄青（偶数行）
    C_PROFIT     = "C6EFCE"   # 薄緑
    C_PROFIT_FG  = "276221"
    C_LOSS       = "FFC7CE"   # 薄赤
    C_LOSS_FG    = "9C0006"
    C_TOTAL_BG   = "FFD966"   # 黄（合計行）
    C_TOTAL_FG   = "000000"
    C_BORDER     = "B8CCE4"

    def hdr_font(sz=11, bold=True, color=C_HEADER_FG):
        return Font(name="Yu Gothic UI", size=sz, bold=bold, color=color)

    def body_font(sz=10, bold=False, color="000000"):
        return Font(name="Yu Gothic UI", size=sz, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border(colors=None):
        c = colors or C_BORDER
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def right():
        return Alignment(horizontal="right", vertical="center")

    def left():
        return Alignment(horizontal="left", vertical="center")

    NUM_FMT  = '#,##0'
    PCT_FMT  = '0.0%'
    DATE_FMT = 'YYYY/MM/DD'

    # ════════════════════════════════════════════════════════
    # Sheet1: サマリー
    # ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "サマリー"
    ws.sheet_view.showGridLines = False

    # --- タイトル ---
    ws.merge_cells("B2:K2")
    tc = ws["B2"]
    tc.value = f"バックテスト結果  {start} ～ {end}（約5年）"
    tc.font      = Font(name="Yu Gothic UI", size=16, bold=True, color=C_HEADER_FG)
    tc.fill      = fill(C_HEADER_BG)
    tc.alignment = center()
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:K3")
    sc = ws["B3"]
    sc.value = (f"初期資金 {INITIAL_CASH:,}円/銘柄 × {len(WATCH_LIST)}銘柄"
                f"  コスト上限{MAX_COST_RATIO*100:.0f}%  "
                f"リスク{RISK_PER_TRADE*100:.0f}%  "
                f"ストップATR×{ATR_STOP_MULT}")
    sc.font      = Font(name="Yu Gothic UI", size=10, color="FFFFFF")
    sc.fill      = fill(C_SUB_BG)
    sc.alignment = center()
    ws.row_dimensions[3].height = 18

    # --- 銘柄別テーブル ヘッダー ---
    COL_HDR = ["銘柄名", "コード", "取引数", "勝率", "総損益(円)",
               "収益率", "最大利益(円)", "最大損失(円)", "PF", "平均保有日数"]
    START_ROW = 5
    for ci, h in enumerate(COL_HDR, start=2):
        cell = ws.cell(row=START_ROW, column=ci, value=h)
        cell.font      = hdr_font(11)
        cell.fill      = fill(C_HEADER_BG)
        cell.alignment = center()
        cell.border    = thin_border("FFFFFF")
    ws.row_dimensions[START_ROW].height = 22

    # --- 銘柄別データ行 ---
    for ri, r in enumerate(results):
        row_num = START_ROW + 1 + ri
        td      = r["trades_detail"]
        mxw     = max((t["pnl"] for t in td), default=0)
        mxl     = min((t["pnl"] for t in td), default=0)
        avg_h   = (sum((t["exit_dt"] - t["entry_dt"]).days for t in td) / len(td)
                   if td else 0)
        pf_val  = None if r["profit_factor"] == float("inf") else r["profit_factor"]

        row_data = [
            r["name"], r["symbol"], r["trades"],
            r["win_rate"] / 100,
            r["total_pnl"], r["return_pct"] / 100,
            mxw, mxl, pf_val, round(avg_h, 1),
        ]
        is_profit = r["total_pnl"] >= 0
        bg = C_PROFIT if is_profit else C_LOSS
        fg = C_PROFIT_FG if is_profit else C_LOSS_FG
        alt_bg = C_ALT if ri % 2 == 1 else "FFFFFF"

        for ci, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = thin_border()

            # 損益列(F=col6)と収益率列(G=col7)だけ色分け、他は交互色
            if ci == 6:   # 総損益
                cell.fill = fill(bg)
                cell.font = body_font(color=fg, bold=True)
                cell.number_format = '#,##0;[Red]-#,##0'
                cell.alignment = right()
            elif ci == 7:  # 収益率
                cell.fill = fill(bg)
                cell.font = body_font(color=fg, bold=True)
                cell.number_format = '+0.0%;-0.0%'
                cell.alignment = right()
            elif ci in (8, 9):   # 最大利益/最大損失
                cell.fill = fill(alt_bg)
                cell.font = body_font()
                cell.number_format = '#,##0;[Red]-#,##0'
                cell.alignment = right()
            elif ci == 4:  # 勝率
                cell.fill = fill(alt_bg)
                cell.font = body_font()
                cell.number_format = '0.0%'
                cell.alignment = center()
            elif ci == 10:  # PF
                cell.fill = fill(alt_bg)
                cell.font = body_font()
                cell.number_format = '0.00'
                cell.alignment = center()
            else:
                cell.fill = fill(alt_bg)
                cell.font = body_font()
                cell.alignment = center() if ci in (3, 5) else left()
        ws.row_dimensions[row_num].height = 18

    # --- 合計行 ---
    total_row = START_ROW + 1 + len(results)
    pf_total  = None if m["profit_factor"] == float("inf") else m["profit_factor"]
    total_data = [
        "【合計】", "", m["n_trades"],
        m["win_rate"] / 100,
        total_pnl, total_ret / 100,
        m["best_trade"], m["worst_trade"],
        pf_total, round(m["avg_hold_days"], 1),
    ]
    for ci, val in enumerate(total_data, start=2):
        cell = ws.cell(row=total_row, column=ci, value=val)
        cell.font   = Font(name="Yu Gothic UI", size=10, bold=True, color=C_TOTAL_FG)
        cell.fill   = fill(C_TOTAL_BG)
        cell.border = thin_border("B8860B")
        if ci == 4:
            cell.number_format = '0.0%'; cell.alignment = center()
        elif ci == 6:
            cell.number_format = '#,##0;[Red]-#,##0'; cell.alignment = right()
        elif ci == 7:
            cell.number_format = '+0.0%;-0.0%'; cell.alignment = right()
        elif ci in (8, 9):
            cell.number_format = '#,##0;[Red]-#,##0'; cell.alignment = right()
        elif ci == 10:
            cell.number_format = '0.00'; cell.alignment = center()
        else:
            cell.alignment = center() if ci in (3, 5) else left()
    ws.row_dimensions[total_row].height = 20

    # --- 列幅 ---
    col_widths = [2, 16, 10, 6, 7, 14, 8, 14, 14, 7, 12]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # --- ポートフォリオ指標ボックス ---
    sign = "+" if total_pnl >= 0 else ""
    pf_disp = "∞" if m["profit_factor"] == float("inf") else f"{m['profit_factor']:.2f}"
    metrics_data = [
        ("総損益",             f"{sign}{total_pnl:,.0f} 円"),
        ("総収益率",            f"{sign}{total_ret:.2f} %"),
        ("年率換算(CAGR)",      f"{cagr*100:+.2f} %"),
        ("最大ドローダウン",      f"-{m['max_drawdown']:.2f} %"),
        ("総取引数",            f"{m['n_trades']} 回"),
        ("勝率",               f"{m['win_rate']:.1f} %"),
        ("プロフィットファクター", pf_disp),
        ("平均保有日数",         f"{m['avg_hold_days']:.1f} 日"),
        ("平均損益/トレード",     f"{m['avg_pnl']:+,.0f} 円"),
        ("最大利益トレード",      f"+{m['best_trade']:,.0f} 円"),
        ("最大損失トレード",      f"{m['worst_trade']:,.0f} 円"),
        ("決済: シグナル",       f"{m['signal_count']} 回"),
        ("決済: ストップ",       f"{m['stop_count']} 回"),
        ("決済: 強制クローズ",    f"{m['force_count']} 回"),
    ]
    METRIC_START_ROW = total_row + 2
    ws.merge_cells(f"B{METRIC_START_ROW}:C{METRIC_START_ROW}")
    mh = ws.cell(row=METRIC_START_ROW, column=2, value="ポートフォリオ指標")
    mh.font = hdr_font(12); mh.fill = fill(C_HEADER_BG); mh.alignment = center()
    ws.row_dimensions[METRIC_START_ROW].height = 22

    for mi, (key, val) in enumerate(metrics_data):
        r_num = METRIC_START_ROW + 1 + mi
        kc = ws.cell(row=r_num, column=2, value=key)
        vc = ws.cell(row=r_num, column=3, value=val)
        kc.font = body_font(bold=True); kc.fill = fill(C_ALT)
        kc.alignment = left(); kc.border = thin_border()
        vc.font = body_font()
        vc.fill = fill("FFFFFF")
        vc.alignment = right(); vc.border = thin_border()
        ws.row_dimensions[r_num].height = 17

    # ════════════════════════════════════════════════════════
    # Sheet2: 全トレード一覧
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("全トレード")
    ws2.sheet_view.showGridLines = False

    trade_hdrs = ["銘柄名", "コード", "エントリー日", "決済日",
                  "保有日数", "買値(円)", "売値(円)", "株数",
                  "損益(円)", "損益率", "決済理由"]
    for ci, h in enumerate(trade_hdrs, start=2):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = hdr_font(11); cell.fill = fill(C_HEADER_BG)
        cell.alignment = center(); cell.border = thin_border("FFFFFF")
    ws2.row_dimensions[2].height = 22

    reason_map = {"signal": "シグナル", "stop": "ストップ", "force_close": "強制"}
    sorted_trades = sorted(all_trades, key=lambda t: t["entry_dt"])
    for ri, t in enumerate(sorted_trades):
        row_num = 3 + ri
        hold    = (t["exit_dt"] - t["entry_dt"]).days
        pnl_r   = (t["exit_price"] - t["entry_price"]) / t["entry_price"]
        is_p    = t["pnl"] >= 0
        row_bg  = C_ALT if ri % 2 == 1 else "FFFFFF"

        row_vals = [
            t["name"], t["symbol"],
            t["entry_dt"].strftime("%Y/%m/%d"),
            t["exit_dt"].strftime("%Y/%m/%d"),
            hold, t["entry_price"], t["exit_price"], t["qty"],
            t["pnl"], pnl_r,
            reason_map.get(t["reason"], t["reason"]),
        ]
        for ci, val in enumerate(row_vals, start=2):
            cell = ws2.cell(row=row_num, column=ci, value=val)
            cell.border = thin_border()
            cell.font   = body_font()
            if ci == 10:   # 損益
                cell.fill = fill(C_PROFIT if is_p else C_LOSS)
                cell.font = body_font(color=C_PROFIT_FG if is_p else C_LOSS_FG, bold=True)
                cell.number_format = '#,##0;[Red]-#,##0'
                cell.alignment = right()
            elif ci == 11:  # 損益率
                cell.fill = fill(C_PROFIT if is_p else C_LOSS)
                cell.font = body_font(color=C_PROFIT_FG if is_p else C_LOSS_FG)
                cell.number_format = '+0.0%;-0.0%'
                cell.alignment = right()
            elif ci in (7, 8, 9):
                cell.fill = fill(row_bg)
                cell.number_format = '#,##0.0'
                cell.alignment = right()
            elif ci == 5:
                cell.fill = fill(row_bg)
                cell.number_format = '#,##0'
                cell.alignment = right()
            else:
                cell.fill = fill(row_bg)
                cell.alignment = center()
        ws2.row_dimensions[row_num].height = 16

    ws2.auto_filter.ref = f"B2:{get_column_letter(2 + len(trade_hdrs) - 1)}{2 + len(sorted_trades)}"
    t2_col_w = [2, 14, 10, 11, 11, 7, 10, 10, 6, 12, 9, 10]
    for ci, w in enumerate(t2_col_w, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet3: 損益推移（グラフ用データ + グラフ）
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("損益推移")
    ws3.sheet_view.showGridLines = False

    # 累積損益データを日付順に構築
    sorted_td = sorted(all_trades, key=lambda t: t["exit_dt"])
    ws3.cell(row=1, column=1, value="決済日").font   = hdr_font()
    ws3.cell(row=1, column=2, value="累積損益(円)").font = hdr_font()
    ws3.cell(row=1, column=1).fill = fill(C_HEADER_BG)
    ws3.cell(row=1, column=2).fill = fill(C_HEADER_BG)
    ws3.cell(row=1, column=1).alignment = center()
    ws3.cell(row=1, column=2).alignment = center()

    cum = 0
    ws3.cell(row=2, column=1, value=start).number_format = DATE_FMT
    ws3.cell(row=2, column=2, value=0)
    for ri, t in enumerate(sorted_td, start=3):
        cum += t["pnl"]
        dc = ws3.cell(row=ri, column=1, value=t["exit_dt"].strftime("%Y/%m/%d"))
        vc = ws3.cell(row=ri, column=2, value=round(cum))
        dc.alignment = center()
        vc.number_format = '#,##0;[Red]-#,##0'
        vc.alignment = right()

    data_rows = 2 + len(sorted_td)
    ws3.column_dimensions["A"].width = 13
    ws3.column_dimensions["B"].width = 14

    # 折れ線グラフ（累積損益推移）
    line = LineChart()
    line.title  = "累積損益推移"
    line.style  = 10
    line.y_axis.title = "累積損益 (円)"
    line.x_axis.title = "決済日"
    line.width  = 22
    line.height = 14
    line.y_axis.numFmt = '#,##0'

    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=data_rows)
    line.add_data(data_ref, titles_from_data=True)
    line.series[0].graphicalProperties.line.solidFill = "2E75B6"
    line.series[0].graphicalProperties.line.width     = 20000
    cats = Reference(ws3, min_col=1, min_row=2, max_row=data_rows)
    line.set_categories(cats)
    ws3.add_chart(line, "D2")

    # ════════════════════════════════════════════════════════
    # Sheet1 に棒グラフ2本追加（サマリーシートに戻る）
    # ════════════════════════════════════════════════════════
    # グラフ用データ（非表示列 M/N に書き出し）
    GCOL = 13  # 列M
    ws.cell(row=4, column=GCOL,   value="銘柄")
    ws.cell(row=4, column=GCOL+1, value="損益(円)")
    ws.cell(row=4, column=GCOL+2, value="勝率(%)")
    for gi, r in enumerate(results):
        ws.cell(row=5+gi, column=GCOL,   value=r["name"])
        ws.cell(row=5+gi, column=GCOL+1, value=round(r["total_pnl"]))
        ws.cell(row=5+gi, column=GCOL+2, value=round(r["win_rate"], 1))

    n = len(results)

    # 棒グラフ①: 銘柄別損益
    bar1 = BarChart()
    bar1.type    = "col"
    bar1.title   = "銘柄別 総損益"
    bar1.style   = 10
    bar1.y_axis.title = "損益 (円)"
    bar1.y_axis.numFmt = '#,##0'
    bar1.width   = 20
    bar1.height  = 13

    d1 = Reference(ws, min_col=GCOL+1, min_row=4, max_row=4+n)
    c1 = Reference(ws, min_col=GCOL,   min_row=5, max_row=4+n)
    bar1.add_data(d1, titles_from_data=True)
    bar1.set_categories(c1)
    # 利益/損失で色分け
    for idx, r in enumerate(results):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = C_PROFIT_FG if r["total_pnl"] >= 0 else "9C0006"
        bar1.series[0].dPt.append(pt)
    ws.add_chart(bar1, "E6")

    # 棒グラフ②: 銘柄別勝率
    bar2 = BarChart()
    bar2.type    = "col"
    bar2.title   = "銘柄別 勝率"
    bar2.style   = 10
    bar2.y_axis.title = "勝率 (%)"
    bar2.y_axis.numFmt = '0.0'
    bar2.y_axis.scaling.min = 0
    bar2.y_axis.scaling.max = 100
    bar2.width   = 20
    bar2.height  = 13

    d2 = Reference(ws, min_col=GCOL+2, min_row=4, max_row=4+n)
    bar2.add_data(d2, titles_from_data=True)
    bar2.set_categories(c1)
    for idx in range(n):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = "2E75B6"
        bar2.series[0].dPt.append(pt)
    ws.add_chart(bar2, "E22")

    wb.save(path)
    return path


# ── メイン ───────────────────────────────────────────────────
def main():
    end   = datetime.today().strftime("%Y-%m-%d")
    start = (datetime.today() - timedelta(days=365 * 5 + 60)).strftime("%Y-%m-%d")

    print(f"\n{'='*65}")
    print(f"  バックテスト  swing_notify_top10.py 戦略")
    print(f"  期間 : {start} 〜 {end}（約5年）")
    print(f"  資金 : {INITIAL_CASH:,}円/銘柄 × {len(WATCH_LIST)}銘柄"
          f" = {INITIAL_CASH * len(WATCH_LIST):,}円")
    print(f"  設定 : コスト上限{MAX_COST_RATIO*100:.0f}%  "
          f"リスク{RISK_PER_TRADE*100:.0f}%  "
          f"ストップATR×{ATR_STOP_MULT}")
    print(f"{'='*65}")

    results    = []
    all_trades = []

    for symbol, name in WATCH_LIST:
        print(f"  取得中: {name}({symbol}) ...", end="\r", flush=True)
        try:
            df = fetch_data(symbol, start, end)
            df = add_indicators(df)
            df = add_signals(df)
            res = backtest_symbol(df, symbol, name)
            results.append(res)
            all_trades.extend(res["trades_detail"])
        except Exception as e:
            print(f"\n  ⚠ {name}({symbol}) スキップ: {e}")

    print(" " * 60, end="\r")

    m          = portfolio_metrics(all_trades)
    total_pnl  = sum(r["total_pnl"] for r in results)
    total_init = INITIAL_CASH * len(WATCH_LIST)
    total_ret  = total_pnl / total_init * 100
    years      = (datetime.strptime(end, "%Y-%m-%d") -
                  datetime.strptime(start, "%Y-%m-%d")).days / 365.25
    cagr       = ((total_init + total_pnl) / total_init) ** (1 / years) - 1

    W  = 72   # 表の幅
    SB = "═"  # 二重線
    DB = "─"  # 単線

    def bar(ch=DB, w=W): return ch * w
    def row(*cols, widths, align="left"):
        """cols を widths に合わせて整形した1行を返す。"""
        cells = []
        for i, (val, w) in enumerate(zip(cols, widths)):
            if i == 0:
                cells.append(f" {val:<{w}}")
            else:
                cells.append(f"{val:>{w}} ")
        return "│" + "│".join(cells) + "│"

    def pnl_str(v):
        return f"+{v:,.0f}" if v >= 0 else f"{v:,.0f}"

    def pct_str(v):
        return f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%"

    def pf_str(v):
        return "  ∞" if v == float("inf") else f"{v:.2f}"

    # ── ヘッダー ──────────────────────────────────────────────
    print()
    print(f"╔{bar(SB)}╗")
    title = f"バックテスト結果  {start} 〜 {end}  (約5年)"
    print(f"║ {title:<{W-2}} ║")
    init_str = (f"初期資金 {INITIAL_CASH:,}円/銘柄 × {len(WATCH_LIST)}銘柄"
                f" = {total_init:,}円   "
                f"コスト上限{MAX_COST_RATIO*100:.0f}%  "
                f"リスク{RISK_PER_TRADE*100:.0f}%  "
                f"ストップATR×{ATR_STOP_MULT}")
    print(f"║ {init_str:<{W-2}} ║")
    print(f"╠{bar(SB)}╣")

    # ── 銘柄別テーブル ────────────────────────────────────────
    # 列幅: 銘柄名16 | 取引4 | 勝率7 | 損益12 | 収益率8 | 最大利益12 | 最大損失12 | PF6
    CW = [14, 4, 7, 12, 8, 10, 10, 6]
    hdrs = ("銘柄名", "取引", "勝率", "損益(円)", "収益率", "最大利益", "最大損失", "PF")
    print(f"║ {'銘柄別パフォーマンス':<{W-2}} ║")
    print(f"╠{'┬'.join(DB*(w+2) for w in CW)}╣".replace("╠", "╟").replace("╣", "╢"))
    print(row(*hdrs, widths=CW))
    print(f"╠{'┼'.join(DB*(w+2) for w in CW)}╣".replace("╠", "╟").replace("╣", "╢"))

    for r in results:
        td    = r["trades_detail"]
        mxw   = max((t["pnl"] for t in td), default=0)
        mxl   = min((t["pnl"] for t in td), default=0)
        print(row(
            r["name"],
            r["trades"],
            f"{r['win_rate']:.1f}%",
            pnl_str(r["total_pnl"]),
            pct_str(r["return_pct"]),
            f"+{mxw:,.0f}",
            f"{mxl:,.0f}",
            pf_str(r["profit_factor"]),
            widths=CW,
        ))

    # 合計行
    print(f"╠{'┼'.join(DB*(w+2) for w in CW)}╣".replace("╠", "╟").replace("╣", "╢"))
    print(row(
        "【合計】",
        m["n_trades"],
        f"{m['win_rate']:.1f}%",
        pnl_str(total_pnl),
        pct_str(total_ret),
        f"+{m['best_trade']:,.0f}",
        f"{m['worst_trade']:,.0f}",
        pf_str(m["profit_factor"]),
        widths=CW,
    ))
    print(f"╚{'╧'.join(SB*(w+2) for w in CW)}╝".replace("╚", "╙").replace("╝", "╜"))

    # ── ポートフォリオ指標 ────────────────────────────────────
    pf_disp = "∞" if m["profit_factor"] == float("inf") else f"{m['profit_factor']:.2f}"
    sign    = "+" if total_pnl >= 0 else ""

    print()
    print(f"╔{bar(SB)}╗")
    print(f"║ {'ポートフォリオ指標':<{W-2}} ║")
    print(f"╠{bar(SB)}╣")

    # 2列レイアウト
    KW, VW = 16, 16   # キー幅・値幅
    GAP = 4
    def kv(k, v): return f"  {k:<{KW}}: {v:<{VW}}"
    def kv2(k1, v1, k2, v2):
        return f"║ {kv(k1,v1)}{' '*GAP}{kv(k2,v2):<{W - KW - VW - GAP - 6}} ║"

    print(kv2("総損益",          f"{sign}{total_pnl:,.0f} 円",
              "年率(CAGR)",      f"{cagr*100:+.2f} %"))
    print(kv2("総収益率",         f"{sign}{total_ret:.2f} %",
              "最大DD",          f"-{m['max_drawdown']:.2f} %"))
    print(kv2("総取引数",         f"{m['n_trades']} 回",
              "平均保有日数",     f"{m['avg_hold_days']:.1f} 日"))
    print(kv2("勝率",            f"{m['win_rate']:.1f} %",
              "平均損益",        f"{m['avg_pnl']:+,.0f} 円"))
    print(kv2("PF",             pf_disp,
              "最大利益",        f"+{m['best_trade']:,.0f} 円"))
    exit_str = (f"シグナル {m['signal_count']} / "
                f"ストップ {m['stop_count']} / "
                f"強制 {m['force_count']}")
    print(kv2("決済内訳",         exit_str,
              "最大損失",        f"{m['worst_trade']:,.0f} 円"))
    print(f"╚{bar(SB)}╝")
    print()

    # ── Excel 出力 ────────────────────────────────────────────
    xlsx_path = "backtest_result.xlsx"
    export_excel(results, all_trades, m, start, end, xlsx_path)
    abs_path  = os.path.abspath(xlsx_path)
    print(f"  Excel を保存しました: {abs_path}")

    # OS に応じて自動で開く
    try:
        if sys.platform == "win32":
            os.startfile(abs_path)
        elif sys.platform == "darwin":
            subprocess.run(["open", abs_path], check=False)
        else:
            subprocess.run(["xdg-open", abs_path], check=False)
    except Exception:
        pass
    print()


if __name__ == "__main__":
    main()
